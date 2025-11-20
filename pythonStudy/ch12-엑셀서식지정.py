import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

book = openpyxl.load_workbook("./엑셀데이터/2024년_광고비_병합.xlsx")
sheet = book.active

# 서식 지정
font = Font(name="굴림", size=12, bold=True)
alignment_center = Alignment(horizontal="center")
color_orange = PatternFill(start_color="FFD732", end_color="FFD732", fill_type="solid")
color_grey = PatternFill(start_color="D2D2D2", end_color="D2D2D2", fill_type="solid")
border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

# 서식 적용
sheet["A14"].value = "합계"
sheet["A14"].font = font
sheet["A14"].alignment = alignment_center
sheet["A14"].fill = color_orange
sheet["A14"].border = border

# 합계 산출
sheet["B14"].value = "=SUM(B2:B13)"
sheet["C14"].value = "=SUM(C2:C13)"
sheet["D14"].value = "=SUM(D2:D13)"

# 광고비 데이터 부분에 서식 적용
for row in sheet["B2:D13"]:
    for cell in row:
        cell.font = font
        cell.alignment = alignment_center
        cell.fill = color_grey
        cell.border = border

# 합계 부분에 서식 적용
for row in sheet["B14:D14"]:
    for cell in row:
        cell.font = font
        cell.alignment = alignment_center
        cell.fill = color_orange
        cell.border = border

book.save("./엑셀데이터/2024년_광고비_병합_서식변경.xlsx")