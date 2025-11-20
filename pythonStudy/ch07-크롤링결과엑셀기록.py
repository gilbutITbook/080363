import requests
from bs4 import BeautifulSoup
import urllib.request as req
import os
import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

# 엑셀 파일 생성
book = openpyxl.Workbook()
sheet = book.active

code = requests.get("https://www.moviechart.co.kr/rank/realtime/index/image")
soup = BeautifulSoup(code.text, "html.parser")
title_elements = soup.select("h3 > a")
image_elements = soup.select("li.movieBox-item > a > img")
percent_elements = soup.select("li.ticketing > span")
date_elements = soup.select("li.movie-launch")

# 이미지 내려받을 폴더 생성
folder_name = "./무비차트 포스터"
if not os.path.exists(folder_name):
    os.mkdir(folder_name)

num = 1
for title, image, percent, date in zip(title_elements, image_elements, percent_elements, date_elements):
    print(f"제목 : {title.text}")
    print(f"예매율 : {percent.text}")
    print(f"개봉일 : {date.text.strip()[6:]}")
    img_url = "https://www.moviechart.co.kr" + image.attrs["src"]
    req.urlretrieve(img_url, f"{folder_name}/{num}.jpg")
    # 이미지 크기 조절
    img = PILImage.open(f"{folder_name}/{num}.jpg")
    img_resized = img.resize((int(img.width * 0.5), int(img.height * 0.5)))
    img_resized.save(f"{folder_name}/{num}_resized.jpg")
    img.close()
    # 엑셀에 이미지 첨부
    sheet.add_image(Image(f"{folder_name}/{num}_resized.jpg"), f"A{num}")
    # 엑셀 시트에 영화 제목, 예매율, 개봉일 기록
    sheet.cell(row=num, column=2).value = title.text
    sheet.cell(row=num, column=3).value = percent.text
    sheet.cell(row=num, column=4).value = date.text.strip()[6:]
    # 행 높이 조절
    sheet.row_dimensions[num].height = 98
    num += 1
    print("----------------------------")
# 열 너비 조절
sheet.column_dimensions["A"].width = 11
sheet.column_dimensions["B"].width = 35
sheet.column_dimensions["C"].width = 15
sheet.column_dimensions["D"].width = 21
book.save("./무비차트.xlsx")