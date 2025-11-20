import openpyxl
import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pandas import ExcelWriter

df = pd.read_excel("./엑셀데이터/아파트분양가.xlsx")
location_list = df["지역명"].unique()
# print(location_list)
for location in location_list:
    df_location = df[df["지역명"] == location]
    year_list = df_location["연도"].unique()
    writer = ExcelWriter(f"./엑셀데이터/아파트분양가_{location}.xlsx")
    for year in year_list:
        df_year = df_location[df_location["연도"] == year]
        df_year = df_year.sort_values(by="월", ascending=True)
        # print(df_year)
        df_year.to_excel(writer, sheet_name=str(year), index=None)
    writer.close()
    print(f"{location} 지역 엑셀 분리 완료")

# 서식 지정
font = Font(name="맑은 고딕", size=12, bold=True)
alignment_center = Alignment(horizontal="center")
color_orange = PatternFill(start_color="FFD732", end_color="FFD732", patternType="solid")
color_grey = PatternFill(start_color="D2D2D2", end_color="D2D2D2", patternType="solid")
border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

for location in location_list:
    book = openpyxl.load_workbook(f"./엑셀데이터/아파트분양가_{location}.xlsx")
    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        # 열 너비 조절
        sheet.column_dimensions["B"].width = 40
        # 제목 행 서식 적용
        for row in sheet["A1:E1"]:
            for cell in row:
                cell.font = font
                cell.alignment = alignment_center
                cell.fill = color_orange
                cell.border = border
        # 데이터 부분 서식 적용
        for row in sheet["A2:E{}".format(sheet.max_row)]:
            for cell in row:
                cell.font = font
                cell.alignment = alignment_center
                cell.fill = color_grey
                cell.border = border
        chart = BarChart()
        chart.title = f"{location} 지역 {sheet_name} 년도 아파트분양가"
        data = Reference(sheet, range_string=f"{sheet_name}!E1:E{sheet.max_row}")
        chart.add_data(data, titles_from_data=True)
        sheet.add_chart(chart, "F1")
    book.save(f"./엑셀데이터/아파트분양가_{location}.xlsx")
    print(f"{location} 지역 엑셀 작업 완료")
