import datetime
import os
import time
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from selenium import webdriver
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.common.by import By

# 숫자 변환 함수 정의
def string_to_integer(view):
    if "만회" in view:
        view = view.replace("만회", "")
        view = int(float(view) * 10000)
    elif "천회" in view:
        view = view.replace("천회", "")
        view = int(float(view) * 1000)
    elif "회" in view:
        view = view.replace("회", "")
        view = int(view)
    return view

# 유튜브 접속
opt = webdriver.ChromeOptions()
opt.add_experimental_option("detach", True)
browser = webdriver.Chrome(options=opt)
browser.get("https://www.youtube.com/@nadocoding/videos")
time.sleep(3)

# 스크롤 끝까지 내리기
same_num_cnt = 0
prev_height = 0
while True:
    ActionChains(browser).key_down(Keys.PAGE_DOWN).perform()
    cur_height = browser.execute_script("return document.documentElement.scrollTop")
    if prev_height == cur_height:
        same_num_cnt += 1
    else:
        same_num_cnt = 0
    if same_num_cnt == 200:
        break
    prev_height = cur_height

# 크롤링
title_elements = reversed(browser.find_elements(By.CSS_SELECTOR, "#contents #video-title"))
views_elements = reversed(browser.find_elements(By.CSS_SELECTOR, "span.inline-metadata-item.style-scope.ytd-video-meta-block:nth-child(3)"))
date_elements = reversed(browser.find_elements(By.CSS_SELECTOR, "span.inline-metadata-item.style-scope.ytd-video-meta-block:nth-child(4)"))
link_elements = reversed(browser.find_elements(By.CSS_SELECTOR, "a#video-title-link"))

# 엑셀 파일 생성
if not os.path.exists("./유튜브채널조회수"):
    os.mkdir("./유튜브채널조회수")
book = openpyxl.Workbook()
sheet = book.active
sheet.cell(row=1, column=1).value = "영상 제목"
sheet.cell(row=1, column=2).value = "조회수"
sheet.cell(row=1, column=3).value = "업로드 날짜"
sheet.cell(row=1, column=4).value = "영상 링크"
sheet.column_dimensions["A"].width = 85
sheet.column_dimensions["B"].width = 10
sheet.column_dimensions["C"].width = 13
sheet.column_dimensions["D"].width = 50

# 결과 출력
row_num = 2
for view, date, title, link in zip(views_elements, date_elements, title_elements, link_elements):
    view_integer = string_to_integer(view.text.replace("조회수 ", ""))
    video_link = link.get_attribute("href")
    print(f"영상 제목 : {title.text}")
    print(f"조회수 : {view_integer}")
    print(f"업로드 날짜 : {date.text}")
    print(f"링크 : {video_link}")
    print("-------------------------------------")
    sheet.cell(row=row_num, column=1).value = title.text
    sheet.cell(row=row_num, column=2).value = view_integer
    sheet.cell(row=row_num, column=3).value = date.text
    sheet.cell(row=row_num, column=4).value = video_link
    row_num += 1

# 차트 생성
chart = LineChart()
chart.title = "조회수 추이"
chart.style = 13
chart.y_axis.title = "조회수"
chart.x_axis.title = "영상 번호"
data = Reference(sheet, range_string=f"Sheet!B1:B{sheet.max_row}")
chart.add_data(data, titles_from_data=True)
chart.width = 45
chart.height = 25
sheet.add_chart(chart, "E1")

# 서식 지정
header_font = Font(bold=True, size=12)
center_alignment = Alignment(horizontal="center")
header_fill = PatternFill(start_color="D9EAF1", end_color="D9EAF1", patternType="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# 제목 행 서식 적용
for row in sheet["A1:D1"]:
    for cell in row:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill

# 데이터 부분 서식 적용
for row in sheet["B2":f"D{sheet.max_row}"]:
    for cell in row:
        cell.alignment = center_alignment

# 영상 링크 걸기
for row in sheet[f"D2:D{sheet.max_row}"]:
    for cell in row:
        cell.hyperlink = cell.value

# 모든 셀에 서식 적용
for row in sheet.iter_rows():
    for cell in row:
        cell.border = thin_border

today = datetime.datetime.today().strftime("%y%m%d_%H시%M분%S초")
book.save(f"./유튜브채널조회수/{today}.xlsx")
browser.close()
